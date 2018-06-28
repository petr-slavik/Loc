#!/usr/bin/env python3
import os
import shutil
import openpyxl
from colorama import Fore
from colorama import init
init()

#----------------------------- DECLARATION --------------------------
'''
Script will take exported files, cut first 6 characters from name (lcid - for example cs-cz_)
Script needs correctly filled reference.xlsx spreadsheet in a root, 
where in first column is FluxID and second is name of file (it starts on A2 and 2B lines).
Then it creates folders named as FluxID and copy there correct files according to reference table.
Then it will create zip files in 05_toASG folders ready for Flux
'''

path = str(os.path.dirname(os.path.realpath(__file__))) + "\\04_postprep"
temp_path = str(os.path.dirname(os.path.realpath(__file__))) + "\\_temp"


print("\n\n\n-----------------------------------------------------------------------------------------------")
print("<<  Postprocessing Script will prepare zip files ready to Flux according to Reference table  >>")
print("-----------------------------------------------------------------------------------------------\n\n\n")


#-----------------------------  FUNCTIONS  --------------------------
def trunc_6characters():
    '''This cuts language prefix from name of file'''
    for root, dirs, files in os.walk(path):
        for name in files:
            print(Fore.LIGHTYELLOW_EX+"Removing "+name[:6]+" from "+name[6:])
            os.rename(os.path.join(root, name), os.path.join(root, name[6:]))


def empty_folders_delete():
    '''This will delete empty folders '''
    for root, dirs, files in os.walk(path):
        for dir in dirs:
            newDir = os.path.join(root, dir)
            try:
                os.removedirs(newDir) #It removes all non-empty folders
                print(Fore.LIGHTWHITE_EX+"Directory empty, deleting...  " + newDir)
            except:
                pass #Exception is if a folder is not empty. Folder will stay untouched.
                #print("Directory not empty and will not be removed")
                #print(" ")


def sorting_files_according_to_reference():
    '''This function reads column A and B in reference Spreadsheet
    and deletes from FluxID folder files which are not connected to FluxID '''
    # Will load reference spreadsheet
    book = openpyxl.load_workbook('reference.xlsx')
    sheet = book.active

    print(Fore.LIGHTCYAN_EX+'Creating FluxID named folders in _temp folder')
    i = 2 #second row in reference sheet
    for i in range(i,100):
        if sheet.cell(row=i, column=1).value == None:
            break
        try:
            # Will create Folders named as FluxIDs in _temp folder
            new_folder = temp_path + "\\" + str(sheet.cell(row=i, column=1).value)
            # Will copy files from 04_postprep to FluxIDs folder in temp folder
            shutil.copytree(path, new_folder)        
            print("Created folder: "+ str(sheet.cell(row=i, column=1).value))
        except:
            pass
        i += 1
    i -= 1 #last row with some relevant value

    # will remove 04_postprep folder and rename _temp to 04_postprep
    shutil.rmtree('04_postprep', ignore_errors=True)
    os.rename("_temp", "04_postprep")

    # load table containing FluxID(folder name) and FileName
    cells = sheet['A2':'B'+str(i)]

    for c1, c2 in cells:
        tree = os.walk(path+"\\"+(str(c1.value)))
        print(Fore.LIGHTGREEN_EX+"To "+str(c1.value)+"\t belongs "+str(c2.value)+"\t...removing other files")
        for root, dirs, soubory in tree:
            for dir in dirs:
                tree2 = os.walk(path+"\\"+(str(c1.value))+"\\"+dir)
                for root, dirs, soubory in tree2:
                    for soubor in soubory:
                        if str(soubor)[:-8] != str(c2.value).split(".")[0]:
                            #.split(".")[0] on c2.value will remove extension from file if it exists
                            #print("\nMazu soubor: " + soubor)
                            #print("ze slozky: " + str(c1.value))
                            os.remove(path+"\\"+(str(c1.value))+"\\"+dir+"\\"+soubor)


def make_zip_for_flux():
    '''This will add files in 04_postprep to archives and store then in 05_toASG folder '''
    path_ASG = str(os.path.dirname(os.path.realpath(__file__))) + "\\05_toASG"

    for dir in os.listdir(path):
        print(Fore.LIGHTMAGENTA_EX+"Making zip of " + dir)
        shutil.make_archive(path_ASG+"\\"+dir, 'zip', path+"\\"+dir)
    print('\nZips are stored in '+path_ASG)


	
#----------------------------- START POINT --------------------------

trunc_6characters()
sorting_files_according_to_reference()
empty_folders_delete()
make_zip_for_flux()