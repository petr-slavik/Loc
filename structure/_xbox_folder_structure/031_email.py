import os
import openpyxl
import win32com.client as win32

from colorama import Fore
from colorama import init
init()


def Emailer(text, subject, recipient_to, recipient_cc, auto=False):
    '''This will send email using Outlook
    Keyword arguments:
        text -- content of email
        subject -- subject of email
        recipient_to -- email adresses in TO
        recipient_cc -- email addreses in CC
        auto -- if false then email is send automaticaly
                if true then mail dialog is displayed and you must click on send'''
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = recipient_to
    mail.CC = recipient_cc
    mail.Subject = subject
    mail.HtmlBody = text
    if auto:
        mail.Send()
    else:
        mail.Display(True)


def CopyTable():    
    '''This should copy table from Excel and insert into Outlook message
    NOT WORKING'''
    xl = win32.Dispatch('Excel.Application')
    xl.Visible = False
    xlsx = xl.Workbooks.Open(str(os.path.normpath(os.getcwd() + '\\reference.xlsx')))
    myRange = xlsx.Sheets(1).Range("A1:E6")
    myRange.Copy()


def MarketCounter():
    '''This will count files in to_trans folder, if there is no folder then it returns ?? string'''
    if os.path.exists(to_trans_folder):
        return len([name for name in os.listdir(to_trans_folder) if os.path.isfile(os.path.join(to_trans_folder, name))])
    else:
        print("THERE IS NO 02_totrans FOLDER!!!")
        return "??"


def getAdjusted():
    '''This will count adjusted strings from Excel analysis file
    + Total / Adjusted numbers for QA team '''
    try:
        book_analysis = openpyxl.load_workbook('analysis_'+ os.path.basename(os.getcwd()) +'.xlsx')
        book_reference = openpyxl.load_workbook('reference.xlsx')
    except ValueError as error:
        print(ValueError)
        return "???"
    sheet = book_analysis.active 
    sheet_reference = book_reference.active  
    radek = 2 #second row in analysis sheet
    adjusted_list = []
    total_list = []
    for radek in range(radek,100):
        if sheet.cell(row=radek, column=10).value == None:
            break
        adjusted_line = []
        for sloupec in range(2,10):
            adjusted_line.append(str(sheet.cell(row=radek, column=sloupec).value))
        adjusted_trans = (int(adjusted_line[1])*0.2)+(int(adjusted_line[3])*0.3)+(int(adjusted_line[4])*0.6)+(int(adjusted_line[5])*0.6)+(int(adjusted_line[6])*1)+(int(adjusted_line[7])*1)
        total_trans = sum(list(map(int, adjusted_line)))
        total_list.append(total_trans)
        adjusted_list.append(adjusted_trans)
        radek += 1
    sorted_list = sorted(adjusted_list)
    low_adjusted =  round(sorted_list[0])
    high_adjusted = round(sorted_list[-1])
    average_adjusted = int(sum(adjusted_list)/len(adjusted_list))
    average_total = int(sum(total_list)/len(total_list))
    global qa_data, qa_deadline, trans_deadline
    qa_data = f"{MarketCounter()} x {average_total} total / {average_adjusted} adjusted"
    trans_deadline = str(sheet_reference.cell(coordinate="G2").value)
    qa_deadline = str(sheet_reference.cell(coordinate="H2").value)
    return f"{low_adjusted}-{high_adjusted}"




subject = "Xbox HO " + os.path.basename(os.getcwd())
to_trans_folder = os.path.normpath(os.getcwd() + "\\02_totrans")
recipient_to = "XXX@jonckers.com; ZZZ@jonckers.com"
recipient_cc = "YYY@jonckers.com; UUU@jonckers.com"

text = f"""<html><body><div class=WordSection1>
<p>Hello,</p>
<p>New Xbox handoffs:<br>
{to_trans_folder}</p>
<p>{MarketCounter()} markets, ~{getAdjusted()} adjusted <br>
Translation Deadline: <b>{trans_deadline}</b></p>
<p>QA Deadline: {qa_deadline}<br>
QA Info: {qa_data}</p>
<p>Reference: <br>--insert-reference-table---</p>
<p>Thanks,</p>
<p style='margin:0mm;margin-bottom:.0001pt'><b><span lang=IT style='font-size:9.0pt;font-family:\"Verdana\",sans-serif;color:#1E73BE'>Petr SLAVIK&nbsp;</span></b><span lang=IT style='font-size:9.0pt;font-family:\"Verdana\",sans-serif;color:#797979'>|&nbsp;Localization Engineer | JONCKERS | CZ Office </span><span style='font-size:10.0pt;font-family:\"Verdana\",sans-serif;color:gray'>|&nbsp;</span><span style='font-family:\"Verdana\",sans-serif'><a href=\"http://www.jonckers.com/\"><span style='font-size:10.0pt'>www.jonckers.com</span></a></span></p></div></div></body></html>"""

print(f"\n\nPreparing HO e-mail\n-------------------\nSubject: {subject} \nTO: {recipient_to}\nCC: {recipient_cc}\n-------------------\nMarkets: {MarketCounter()}\nAdjusted: {getAdjusted()}\n\n")
print(Fore.LIGHTMAGENTA_EX+"Complete missing data: \nDEADLINE, REFERENCE TABLE")


Emailer(text, subject, recipient_to, recipient_cc)
#CopyTable()
