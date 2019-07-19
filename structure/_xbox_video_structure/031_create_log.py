import os
import openpyxl


def MarketCounter():
    '''This will count files in to_trans folder, if there is no folder then it returns ?? string'''
    if os.path.exists(to_trans_folder):
        return len([name for name in os.listdir(to_trans_folder) if os.path.isfile(os.path.join(to_trans_folder, name))])
    else:
        print("THERE IS NO 02_totrans FOLDER!!!")
        return "??"


def getAdjusted():
    '''This will count adjusted strings from Excel analysis file
    + Total / Adjusted numbers for QA team '''
    try:
        book_analysis = openpyxl.load_workbook('analysis_'+ os.path.basename(os.getcwd()) +'.xlsx')
    except ValueError as error:
        print(ValueError)
        return "???"
    sheet = book_analysis.active 
    radek = 2 #second row in analysis sheet
    adjusted_list = []
    total_list = []
    for radek in range(radek,100):
        if sheet.cell(row=radek, column=10).value == None:
            break
        adjusted_line = []
        for sloupec in range(2,10):
            adjusted_line.append(str(sheet.cell(row=radek, column=sloupec).value))
        adjusted_trans = (int(adjusted_line[1])*0.2)+(int(adjusted_line[2])*0.2)+(int(adjusted_line[3])*0.3)+(int(adjusted_line[4])*0.6)+(int(adjusted_line[5])*0.6)+(int(adjusted_line[6])*1)+(int(adjusted_line[7])*1)
        total_trans = sum(list(map(int, adjusted_line)))
        total_list.append(total_trans)
        adjusted_list.append(adjusted_trans)
        radek += 1
    sorted_list = sorted(adjusted_list)
    low_adjusted =  round(sorted_list[0])
    high_adjusted = round(sorted_list[-1])
    return f"{low_adjusted}-{high_adjusted}"


to_trans_folder = os.path.normpath(os.getcwd() + "\\02_totrans")

text = f"""Hi Tomas,

There are packages for translation:
{to_trans_folder}

{MarketCounter()} markets, ~{getAdjusted()} adjusted

Thanks,"""


def xbox_video_txt():
    with open("text_for_mail.txt", "w") as log:
        log.write(text)
    print(f"\nMarkets: {MarketCounter()}\nAdjusted: {getAdjusted()}\n\n")


def main():
    xbox_video_txt()


if __name__ == '__main__':
    main()
