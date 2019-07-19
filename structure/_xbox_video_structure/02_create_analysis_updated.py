import os
import re
import openpyxl
import codecs
import xml.etree.ElementTree as etree
'''


Notes:
	added automatic sorting for MemoQ stats
	script simplified
	added new column "Agresso Code" and detect code based on the given matrix
	new stats format from ASG is in place
	engineer can select what MemoQ stats type to proceed
	added automatic file name naming based on trados stats info
	sorted by agresso code
	removed xx-xx language code from report
todo:

WARNING:
	
'''
codes = [["eng-GB", "ENG"],["en-GB", "ENG"],["ger-AT", "AUS"],["de-AT", "AUS"],["af-ZA", "AFR"], ["sq-AL", "ALB"], ["am-ET", "AMH"], ["ar-SA", "ARA"], ["hy-AM", "ARM"], ["as-IN", "ASM"], ["az-Latn-AZ", "AZE"], ["bn-BD", "BEB"], ["eu-ES", "EUS"], ["be-BY", "BEL"], ["bn-IN", "BEN"], ["bs-Latn-BA", "BOS"], ["bg-BG", "BUL"], ["ca-ES", "CAT"], ["zh-CN", "CHS"], ["zh-TW", "CHT"], ["zh-HK", "HKC"], ["hr-HR", "CRO"], ["cs-CZ", "CZE"], ["da-dk", "DAN"], ["prs-AF", "PRS"], ["nl-NL", "DUT"], ["en-GB", "UKE"], ["et-EE", "EST"], ["fil-PH", "FIL"], ["fi-FI", "FIN"], ["fr-CA", "CFR"], ["fr-FR", "FRE"], ["gl-ES", "GLG"], ["ka-GE", "GEO"], ["de-DE", "GER"], ["el-GR", "GRE"], ["gu-IN", "GUJ"], ["ha-Latn-NG", "HAU"], ["he-IL", "HEB"], ["hi-IN", "HIN"], ["hu-HU", "HUN"], ["is-IS", "ICE"], ["ig-NG", "IBO"], ["id-ID", "IND"], ["ga-IE", "GLE"], ["xh-ZA", "XHO"], ["zu-ZA", "ZUL"], ["it-IT", "ITA"], ["ja-JP", "JPN"], ["kn-IN", "KAN"], ["kk-KZ", "KAZ"], ["km-KH", "KHM"], ["quc-Latn-GT", "QUC"], ["rw-RW", "KIN"], ["sw-KE", "SWA"], ["kok-IN", "KOK"], ["ko-KR", "KOR"], ["ku-ARAB-IQ", "KUR"], ["ky-KG", "KIR"], ["lo-LA", "LAO"], ["lv-LV", "LAV"], ["lt-LT", "LIT"], ["lb-LU", "LUX"], ["mk-MK", "MAC"], ["ms-MY", "MLY"], ["ml-IN", "MAL"], ["mt-MT", "MLT"], ["mi-NZ", "MRI"], ["mr-IN", "MAR"], ["mn-MN", "MON"], ["ne-NP", "NEP"], ["nb-NO", "NOR"], ["nn-NO", "NNO"], ["or-IN", "ORI"], ["fa-IR", "PER"], ["pl-PL", "POL"], ["pt-BR", "BPO"], ["pt-PT", "POR"], ["pa-IN", "PUN"], ["pa-Arab-PK", "PAN"], ["quz-PE", "QUP"], ["ro-RO", "RUM"], ["ru-RU", "RUS"], ["gd-GB", "GLA"], ["sr-Cyrl-BA", "SRB"], ["sr-Cyrl-RS", "SRC"], ["sr-Latn-RS", "SRL"], ["nso-ZA", "SOT"], ["tn-ZA", "TSN"], ["sd-Arab-PK", "SND"], ["si-LK", "SIN"], ["sk-SK", "SLK"], ["sl-SI", "SLV"], ["es-MX", "MEX"], ["es-ES", "SPA"], ["sv-SE", "SWE"], ["tg-Cyrl-TJ", "TGK"], ["ta-IN", "TAM"], ["tt-RU", "TAT"], ["te-IN", "TEL"], ["th-TH", "THA"], ["ti-ET", "TIR"], ["tr-TR", "TUR"], ["tk-TM", "TUK"], ["uk-UA", "UKR"], ["ur-PK", "URD"], ["ug-CN", "UIG"], ["uz-Latn-UZ", "UZB"], ["ca-ES-valencia", "VAL"], ["vi-VN", "VIE"], ["cy-GB", "WEL"], ["wo-SN", "WOL"], ["yo-NG", "YOR"]]
def getAgressoCode(langCode):
	returnValue="n/a"
	for code in codes:
		if code[0].lower()==langCode:
			returnValue=code[1].upper()
			break
	return returnValue

memoq_codes = [["ara-SA", "ar-SA"], ["bul", "bg-BG"], ["cze", "cs-CZ"], ["dan", "da-DK"], ["dut-NL", "nl-NL"], ["est", "et-EE"], ["fin", "fi-FI"], ["fre-CA", "fr-CA"], ["fre-FR", "fr-FR"], ["ger-DE", "de-DE"], ["gre", "el-GR"], ["heb", "he-IL"], ["hin", "hi-IN"], ["hrv", "hr-HR"], ["hun", "hu-HU"], ["ind", "id-ID"], ["ita-IT", "it-IT"], ["jpn", "ja-JP"], ["kaz", "kk-KZ"], ["kor", "ko-KR"], ["lav", "lv-LV"], ["lit", "lt-LT"], ["nnb", "nb-NO"], ["pol", "pl-PL"], ["por-BR", "pt-BR"], ["por-PT", "pt-PT"], ["rum", "ro-RO"], ["rus", "ru-RU"], ["Serbian", "sr-Latn-RS"], ["slo", "sk-SK"], ["spa-ES", "es-ES"], ["spa-MX", "es-MX"], ["swe-SE", "sv-SE"], ["tha", "th-TH"], ["tur", "tr-TR"], ["ukr", "uk-UA"], ["vie", "vi-VN"], ["zho-CN", "zh-CN"], ["zho-TW", "zh-TW"]]
def getLngCode(langCode):
	returnValue=langCode
	for code in memoq_codes:
		if code[0].lower()==langCode:
			returnValue=code[1].lower()
			break
	return returnValue

def writeHeader(activeSheet):
	activeSheet.cell(row = 1, column = 1).value = ""
	activeSheet.cell(row = 1, column = 2).value = "Context match"
	activeSheet.cell(row = 1, column = 3).value = "Repetitions"
	activeSheet.cell(row = 1, column = 4).value = "100%"
	activeSheet.cell(row = 1, column = 5).value = "95% - 99%"
	activeSheet.cell(row = 1, column = 6).value = "85% - 94%"
	activeSheet.cell(row = 1, column = 7).value = "75% - 84%"
	activeSheet.cell(row = 1, column = 8).value = "50% - 74%"
	activeSheet.cell(row = 1, column = 9).value = "No Match"
	activeSheet.cell(row = 1, column = 10).value = "Total"
	activeSheet.cell(row = 1, column = 11).value = "Adjusted for trans"
	return

def updateCellNumber(activeSheet, row, column, value):
	activeSheet[row+str(column)] = int(value)
	return

def updateCellString(activeSheet, row, column, value):
	activeSheet[row+str(column)] = str(value)
	return


def proceedTradosstat(file, groups):
	single_row = ["","",0,0,0,0,0,0,0,0,0,0]
	rowCnt=1
	for group in groups:
		rowCnt+=1
		
		single_row[0] = str(file.lower())[20:25]
		single_row[1] = str(file.lower())[20:25]
		single_row[2] = int(group.find("inContextExact").get("words"))
		single_row[3] = int(group.find("repeated").get("words"))+int(group.find("crossFileRepeated").get("words"))
		single_row[4] = int(group.find("exact").get("words"))
		single_row[5] = int(group.find("fuzzy[@min='95']").get("words"))+int(group.find("internalFuzzy[@min='95']").get("words"))
		single_row[6] = int(group.find("fuzzy[@min='85']").get("words"))+int(group.find("internalFuzzy[@min='85']").get("words"))
		single_row[7] = int(group.find("fuzzy[@min='75']").get("words"))+int(group.find("internalFuzzy[@min='75']").get("words"))
		single_row[8] = int(group.find("fuzzy[@min='50']").get("words"))+int(group.find("internalFuzzy[@min='50']").get("words"))
		single_row[9] = int(group.find("new").get("words"))
	return single_row

def storeStat(aSheet, exported_row, row):
	updateCellString(aSheet, "A", row, exported_row[1])
	updateCellNumber(aSheet, "B", row, exported_row[2])
	updateCellNumber(aSheet, "C", row, exported_row[3])
	updateCellNumber(aSheet, "D", row, exported_row[4])
	updateCellNumber(aSheet, "E", row, exported_row[5])
	updateCellNumber(aSheet, "F", row, exported_row[6])
	updateCellNumber(aSheet, "G", row, exported_row[7])
	updateCellNumber(aSheet, "H", row, exported_row[8])
	updateCellNumber(aSheet, "I", row, exported_row[9])
	updateCellString(aSheet, "J", row, "=SUM(B"+str(row)+":I"+str(row)+")")
	updateCellString(aSheet, "K", row, "=(C"+str(row)+"*0.2)+(D"+str(row)+"*0.2)+(E"+str(row)+"*0.3)+(F"+str(row)+"*0.6)+(G"+str(row)+"*0.6)+(H"+str(row)+"*1)+(I"+str(row)+"*1)")
	
	return

#clear screen
os.system('cls')

#create analysis workbook
wb = openpyxl.Workbook()
#get active sheet and rename it
sheet = wb.active
sheet.title = 'TS2015'
#write headers
aSheet = wb.get_sheet_by_name('TS2015')
writeHeader(aSheet)

#Trados part

#activate sheet
extraFileName = ""
aSheet = wb.get_sheet_by_name('TS2015')
exported_rows=[]
#Go through all subdirs (just to get correct path to trados folder)
for subdir_d, dirs_d, files_d in os.walk(os.getcwd()):
	for dir_d in dirs_d:
		if (dir_d.find("trados")>=0):
			#go through all files in trados subdir
			for subdir, dirs, files in os.walk(subdir_d + os.sep + dir_d):
				for file in files:
					filepath = subdir + os.sep + file
					#if it is analyse file then proceed
					if filepath.endswith(".xml") and (file.startswith("Analyze Files")):
						parser = etree.XMLParser()
						tree = etree.parse(filepath, parser) 
						root = tree.getroot()
						groups = root.findall(".//batchTotal//analyse")
						p = proceedTradosstat(file, groups)
						if p!=[]:
							exported_rows.append(p)
						#get the project name from trados analyse file
						projectNames = root.findall(".//taskInfo")
						for projectName in projectNames:
							extraFileName = "_"+projectName.find("project").get("name")
#sort exported data
exported_rows.sort(key=lambda x: x[1])
#store exported data
row = 2
for exported_row in exported_rows:
	storeStat(aSheet, exported_row, row)
	row=row+1

#save analysis
wb.save('analysis'+extraFileName+'.xlsx')

